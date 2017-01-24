__author__ = 'Rashed'
import requests

from bs4 import BeautifulSoup
val = requests.get('http://www.wscraper.com/')
# print(val.text)

soup = BeautifulSoup(val.text, "html.parser")
print (val.text)
all_value = (soup.find("ul", {"id": "mainMenuNav"}))
#
# all_li = all_value.findAll('li')
# print(all_li)

# for x in all_li:
#     # print(x)
#     print(x.text.upper())


value = soup.find("p", attrs={'class':'lplh-42'})
h2 = value.findAll('span')[0]
h3 = value.findAll('span')[1]
print(h2.text)
print(h3.text)
print(len(value.findAll('span')))

# __author__ = 'Azad'
# from openpyxl import Workbook
# wb = Workbook()
# #
# # female = (('Name', 'place','fairness'),
# #           ('5 jon', 'lalbag','good'),
# #           ('sahaMima','buet','best'),
# #           ('sefali','buet','best'))
#
# # grab the active worksheet
# ws = wb.active
# ws.append([x.text for x in [h2,h3]])
# wb.save("spandas_simple.xlsx")
from openpyxl import Workbook
wb = Workbook()

# grab the active worksheet
ws = wb.active

ws['G1'] =  h2.text
wb.save("save.xlsx")